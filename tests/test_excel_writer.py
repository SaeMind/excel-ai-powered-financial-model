"""Tests for Excel writer output."""

from pathlib import Path

from openpyxl import load_workbook

from src.excel_writer import ExcelFinancialModel
from src.model import FinancialModel
from src.scenarios import create_base_case


def test_excel_writer_contains_formulas(tmp_path: Path) -> None:
    """Workbook should contain editable assumptions and formulas."""
    model = FinancialModel(create_base_case())
    model.run()
    path = tmp_path / "model.xlsx"
    ExcelFinancialModel(model, path).write_model()
    workbook = load_workbook(path, data_only=False)
    assert "Assumptions" in workbook.sheetnames
    assert workbook["P&L Projection"]["B6"].value == "=B4*B5*12"
    assert workbook["Dashboard"]["A4"].value == "='P&L Projection'!D6"
