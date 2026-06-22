"""Excel writer for the health-tech financial model."""

from __future__ import annotations

from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from .model import FinancialModel
from .scenarios import run_scenario_analysis


class ExcelFinancialModel:
    """Create an interactive formula-driven Excel financial model."""

    def __init__(self, model: FinancialModel, output_path: Path) -> None:
        """Initialize workbook writer.

        Parameters:
            model: FinancialModel used to seed default values.
            output_path: Destination XLSX path.
        """
        self.model = model
        self.output_path = output_path
        self.workbook = Workbook()
        self.workbook.remove(self.workbook.active)
        self.dark_fill = PatternFill("solid", fgColor="1F2937")
        self.section_fill = PatternFill("solid", fgColor="E5E7EB")
        self.input_fill = PatternFill("solid", fgColor="FFF2CC")
        self.blue_font = Font(color="0000FF")
        self.black_font = Font(color="000000")
        self.green_font = Font(color="008000")
        self.white_bold = Font(color="FFFFFF", bold=True)
        self.thin_gray = Side(style="thin", color="D1D5DB")
        self.border = Border(bottom=self.thin_gray)
        self.money_fmt = '$#,##0;[Red]($#,##0);-'
        self.money_one_decimal_fmt = '$#,##0.0;[Red]($#,##0.0);-'
        self.percent_fmt = '0.0%;[Red](0.0%);-'
        self.multiple_fmt = '0.0x;[Red](0.0x);-'
        self.number_fmt = '#,##0;[Red](#,##0);-'

    def write_model(self) -> Path:
        """Write and save the complete Excel model.

        Returns:
            Path to saved workbook.
        """
        self.output_path.parent.mkdir(parents=True, exist_ok=True)
        self._create_assumptions_sheet()
        self._create_model_sheet()
        self._create_saas_metrics_sheet()
        self._create_scenario_sheet()
        self._create_dashboard_sheet()
        self._style_workbook()
        self.workbook.calculation.fullCalcOnLoad = True
        self.workbook.calculation.forceFullCalc = True
        self.workbook.save(self.output_path)
        return self.output_path

    def _create_assumptions_sheet(self) -> None:
        """Create editable assumptions sheet with source comments."""
        ws = self.workbook.create_sheet("Assumptions")
        assumptions = self.model.assumptions
        ws["A1"] = "Health-Tech Financial Model Assumptions"
        ws["A1"].font = Font(bold=True, size=16, color="111827")
        ws.merge_cells("A1:D1")
        ws["A3"] = "Parameter"
        ws["B3"] = "Value"
        ws["C3"] = "Unit"
        ws["D3"] = "Notes"
        rows = [
            ("Year 1 Users", assumptions.year1_users, "users", "Seeded model input; replace with traction or forecast source."),
            ("Year 1 ARPU", assumptions.year1_arpu, "$/user/month", "Average monthly recurring revenue per monetized user."),
            ("Y1 to Y2 User Growth", assumptions.user_growth_y1_y2, "x", "User growth multiplier."),
            ("Y2 to Y3 User Growth", assumptions.user_growth_y2_y3, "x", "User growth multiplier."),
            ("Annual ARPU Growth", assumptions.arpu_growth, "%", "Pricing, expansion, and product mix uplift."),
            ("COGS % of Revenue", assumptions.cogs_percent_revenue, "%", "Cloud, model inference, support, and hosting load."),
            ("Year 1 OpEx", assumptions.opex_base, "$", "Team, product, compliance, marketing, and G&A."),
            ("Annual OpEx Growth", assumptions.opex_growth, "%", "Operating scale growth."),
            ("Tax Rate", assumptions.tax_rate, "%", "Corporate tax assumption."),
            ("Discount Rate", assumptions.discount_rate, "%", "Risk-adjusted discount rate."),
            ("CAC", assumptions.cac, "$/new user", "Customer acquisition cost."),
            ("Monthly Churn", assumptions.monthly_churn, "%", "Monthly logo churn."),
            ("Gross Revenue Retention", assumptions.gross_revenue_retention, "%", "Revenue retained before expansion."),
            ("Net Revenue Retention", assumptions.net_revenue_retention, "%", "Revenue retained after expansion."),
            ("Starting Cash", assumptions.starting_cash, "$", "Available capital before Year 1."),
        ]
        for row_idx, row in enumerate(rows, start=4):
            for col_idx, value in enumerate(row, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)
            ws.cell(row=row_idx, column=2).font = self.blue_font
            ws.cell(row=row_idx, column=2).fill = self.input_fill
            ws.cell(row=row_idx, column=2).comment = Comment(row[3], "ChatGPT")
        self._format_header(ws, 3, 1, 4)
        for row in [6, 7, 15]:
            ws.cell(row=row, column=2).number_format = self.multiple_fmt
        for row in [8, 9, 10, 11, 15, 16, 17]:
            ws.cell(row=row, column=2).number_format = self.percent_fmt
        for row in [5, 12, 14, 18]:
            ws.cell(row=row, column=2).number_format = self.money_fmt
        ws.cell(row=4, column=2).number_format = self.number_fmt
        ws.freeze_panes = "A4"

    def _create_model_sheet(self) -> None:
        """Create formula-driven P&L projection sheet."""
        ws = self.workbook.create_sheet("P&L Projection")
        ws["A1"] = "3-Year P&L Projection ($)"
        ws["A1"].font = Font(bold=True, size=16)
        ws.merge_cells("A1:D1")
        headers = ["Metric", "Year 1", "Year 2", "Year 3"]
        for col, header in enumerate(headers, start=1):
            ws.cell(3, col, header)
        self._format_header(ws, 3, 1, 4)
        rows = [
            ("Users", "=Assumptions!$B$4", "=B4*Assumptions!$B$6", "=C4*Assumptions!$B$7", self.number_fmt),
            ("ARPU / Month", "=Assumptions!$B$5", "=B5*(1+Assumptions!$B$8)", "=C5*(1+Assumptions!$B$8)", self.money_one_decimal_fmt),
            ("Revenue", "=B4*B5*12", "=C4*C5*12", "=D4*D5*12", self.money_fmt),
            ("COGS", "=B6*Assumptions!$B$9", "=C6*Assumptions!$B$9", "=D6*Assumptions!$B$9", self.money_fmt),
            ("Gross Profit", "=B6-B7", "=C6-C7", "=D6-D7", self.money_fmt),
            ("Gross Margin", "=IFERROR(B8/B6,0)", "=IFERROR(C8/C6,0)", "=IFERROR(D8/D6,0)", self.percent_fmt),
            ("OpEx", "=Assumptions!$B$10", "=B10*(1+Assumptions!$B$11)", "=C10*(1+Assumptions!$B$11)", self.money_fmt),
            ("EBITDA", "=B8-B10", "=C8-C10", "=D8-D10", self.money_fmt),
            ("EBITDA Margin", "=IFERROR(B11/B6,0)", "=IFERROR(C11/C6,0)", "=IFERROR(D11/D6,0)", self.percent_fmt),
            ("Tax", "=MAX(0,B11*Assumptions!$B$12)", "=MAX(0,C11*Assumptions!$B$12)", "=MAX(0,D11*Assumptions!$B$12)", self.money_fmt),
            ("Net Income", "=B11-B13", "=C11-C13", "=D11-D13", self.money_fmt),
            ("Net Margin", "=IFERROR(B14/B6,0)", "=IFERROR(C14/C6,0)", "=IFERROR(D14/D6,0)", self.percent_fmt),
            ("Ending Cash", "=Assumptions!$B$18+B14", "=B16+C14", "=C16+D14", self.money_fmt),
        ]
        for row_idx, row in enumerate(rows, start=4):
            ws.cell(row_idx, 1, row[0])
            for col_idx in range(2, 5):
                ws.cell(row_idx, col_idx, row[col_idx - 1])
                ws.cell(row_idx, col_idx).number_format = row[4]
                ws.cell(row_idx, col_idx).font = self.green_font
        self._apply_total_style(ws, [8, 11, 14, 16], 1, 4)
        ws.freeze_panes = "B4"

    def _create_saas_metrics_sheet(self) -> None:
        """Create SaaS metrics sheet with formula-driven unit economics."""
        ws = self.workbook.create_sheet("SaaS Metrics")
        ws["A1"] = "Healthcare SaaS Unit Economics"
        ws["A1"].font = Font(bold=True, size=16)
        ws.merge_cells("A1:D1")
        for col, header in enumerate(["Metric", "Year 1", "Year 2", "Year 3"], 1):
            ws.cell(3, col, header)
        self._format_header(ws, 3, 1, 4)
        rows = [
            ("New Users", "='P&L Projection'!B4", "='P&L Projection'!C4-'P&L Projection'!B4", "='P&L Projection'!D4-'P&L Projection'!C4", self.number_fmt),
            ("Implied S&M Spend", "=B4*Assumptions!$B$14", "=C4*Assumptions!$B$14", "=D4*Assumptions!$B$14", self.money_fmt),
            ("LTV", "=('P&L Projection'!B5*12*'P&L Projection'!B9)/MAX(Assumptions!$B$15*12,0.0001)", "=('P&L Projection'!C5*12*'P&L Projection'!C9)/MAX(Assumptions!$B$15*12,0.0001)", "=('P&L Projection'!D5*12*'P&L Projection'!D9)/MAX(Assumptions!$B$15*12,0.0001)", self.money_fmt),
            ("LTV:CAC", "=B6/Assumptions!$B$14", "=C6/Assumptions!$B$14", "=D6/Assumptions!$B$14", self.multiple_fmt),
            ("CAC Payback", "=Assumptions!$B$14/MAX('P&L Projection'!B5*'P&L Projection'!B9,0.0001)", "=Assumptions!$B$14/MAX('P&L Projection'!C5*'P&L Projection'!C9,0.0001)", "=Assumptions!$B$14/MAX('P&L Projection'!D5*'P&L Projection'!D9,0.0001)", '0.0 "mo";-0.0 "mo";-'),
            ("Gross Revenue Retention", "=Assumptions!$B$16", "=Assumptions!$B$16", "=Assumptions!$B$16", self.percent_fmt),
            ("Net Revenue Retention", "=Assumptions!$B$17", "=Assumptions!$B$17", "=Assumptions!$B$17", self.percent_fmt),
            ("Burn Multiple", "=IFERROR(ABS(MIN('P&L Projection'!B14,0))/'P&L Projection'!B6,0)", "=IFERROR(ABS(MIN('P&L Projection'!C14,0))/('P&L Projection'!C6-'P&L Projection'!B6),0)", "=IFERROR(ABS(MIN('P&L Projection'!D14,0))/('P&L Projection'!D6-'P&L Projection'!C6),0)", self.multiple_fmt),
        ]
        for row_idx, row in enumerate(rows, start=4):
            ws.cell(row_idx, 1, row[0])
            for col_idx in range(2, 5):
                ws.cell(row_idx, col_idx, row[col_idx - 1])
                ws.cell(row_idx, col_idx).number_format = row[4]
                ws.cell(row_idx, col_idx).font = self.green_font
        self._apply_total_style(ws, [7, 8, 11], 1, 4)

    def _create_scenario_sheet(self) -> None:
        """Create scenario comparison sheet."""
        ws = self.workbook.create_sheet("Scenario Comparison")
        ws["A1"] = "Scenario Comparison"
        ws["A1"].font = Font(bold=True, size=16)
        ws.merge_cells("A1:D1")
        scenario_results = run_scenario_analysis()
        headers = ["Metric", "Base", "Pessimistic", "Optimistic"]
        for col, header in enumerate(headers, 1):
            ws.cell(3, col, header)
        self._format_header(ws, 3, 1, 4)
        metric_rows = [
            ("Year 3 Users", "year3_users", self.number_fmt),
            ("Year 3 Revenue", "year3_revenue", self.money_fmt),
            ("Year 3 EBITDA", "year3_ebitda", self.money_fmt),
            ("Year 3 EBITDA Margin", "year3_ebitda_margin", self.percent_fmt),
            ("Year 3 Net Income", "year3_net_income", self.money_fmt),
            ("3-Year NPV", "npv_3_year", self.money_fmt),
            ("LTV:CAC", "ltv_cac_ratio_y3", self.multiple_fmt),
            ("CAC Payback", "payback_months_y3", '0.0 "mo";-0.0 "mo";-'),
            ("Ending Cash", "ending_cash_y3", self.money_fmt),
        ]
        scenario_keys = ["base", "pessimistic", "optimistic"]
        for row_idx, (label, key, fmt) in enumerate(metric_rows, start=4):
            ws.cell(row_idx, 1, label)
            for col_idx, scenario_key in enumerate(scenario_keys, start=2):
                ws.cell(row_idx, col_idx, scenario_results[scenario_key][key])
                ws.cell(row_idx, col_idx).number_format = fmt
        self._apply_total_style(ws, [9, 12], 1, 4)

    def _create_dashboard_sheet(self) -> None:
        """Create executive dashboard with KPI cards and charts."""
        ws = self.workbook.create_sheet("Dashboard", 0)
        ws["A1"] = "Executive Dashboard"
        ws["A1"].font = Font(bold=True, size=18)
        ws.merge_cells("A1:H1")
        kpis = [
            ("Year 3 Revenue", "='P&L Projection'!D6", self.money_fmt),
            ("Year 3 EBITDA", "='P&L Projection'!D11", self.money_fmt),
            ("Year 3 EBITDA Margin", "='P&L Projection'!D12", self.percent_fmt),
            ("3-Year NPV", "=NPV(Assumptions!$B$13,'P&L Projection'!B14:D14)", self.money_fmt),
            ("Year 3 Users", "='P&L Projection'!D4", self.number_fmt),
            ("LTV:CAC", "='SaaS Metrics'!D7", self.multiple_fmt),
            ("CAC Payback", "='SaaS Metrics'!D8", '0.0 "mo";-0.0 "mo";-'),
            ("Ending Cash", "='P&L Projection'!D16", self.money_fmt),
        ]
        positions = [(3, 1), (3, 3), (3, 5), (3, 7), (6, 1), (6, 3), (6, 5), (6, 7)]
        for (label, formula, fmt), (row, col) in zip(kpis, positions):
            ws.cell(row, col, label)
            ws.cell(row, col).font = self.white_bold
            ws.cell(row, col).fill = self.dark_fill
            ws.cell(row + 1, col, formula)
            ws.cell(row + 1, col).number_format = fmt
            ws.cell(row + 1, col).font = Font(bold=True, size=13, color="008000")
            ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 1)
            ws.merge_cells(start_row=row + 1, start_column=col, end_row=row + 1, end_column=col + 1)
        self._add_dashboard_charts(ws)

    def _add_dashboard_charts(self, ws) -> None:  # type: ignore[no-untyped-def]
        """Add dashboard charts.

        Parameters:
            ws: Dashboard worksheet.
        """
        pnl = self.workbook["P&L Projection"]
        chart = LineChart()
        chart.title = "Revenue and EBITDA Growth"
        chart.y_axis.title = "Dollars"
        chart.x_axis.title = "Year"
        data = Reference(pnl, min_col=2, max_col=4, min_row=6, max_row=11)
        cats = Reference(pnl, min_col=2, max_col=4, min_row=3, max_row=3)
        chart.add_data(data, from_rows=True, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 7
        chart.width = 15
        ws.add_chart(chart, "A10")

        scenarios = self.workbook["Scenario Comparison"]
        bar = BarChart()
        bar.title = "3-Year NPV by Scenario"
        bar.y_axis.title = "NPV"
        scenario_data = Reference(scenarios, min_col=2, max_col=4, min_row=9, max_row=9)
        scenario_cats = Reference(scenarios, min_col=2, max_col=4, min_row=3, max_row=3)
        bar.add_data(scenario_data, from_rows=True, titles_from_data=False)
        bar.set_categories(scenario_cats)
        bar.dataLabels = DataLabelList()
        bar.dataLabels.showVal = True
        bar.height = 7
        bar.width = 15
        ws.add_chart(bar, "E10")

    def _style_workbook(self) -> None:
        """Apply workbook-wide formatting."""
        for ws in self.workbook.worksheets:
            ws.sheet_view.showGridLines = False
            for row in ws.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(vertical="center")
                    if cell.value is not None:
                        cell.border = self.border
            widths = {"A": 28, "B": 18, "C": 18, "D": 18, "E": 18, "F": 18, "G": 18, "H": 18}
            for col, width in widths.items():
                ws.column_dimensions[col].width = width
        self.workbook.active = 0

    def _format_header(self, ws, row: int, start_col: int, end_col: int) -> None:  # type: ignore[no-untyped-def]
        """Format a row as a section header.

        Parameters:
            ws: Worksheet to format.
            row: Row number.
            start_col: First column number.
            end_col: Last column number.
        """
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = self.white_bold
            cell.fill = self.dark_fill
            cell.alignment = Alignment(horizontal="center")

    def _apply_total_style(self, ws, rows: Iterable[int], start_col: int, end_col: int) -> None:  # type: ignore[no-untyped-def]
        """Apply financial-model total row style.

        Parameters:
            ws: Worksheet to format.
            rows: Row numbers to style.
            start_col: First column number.
            end_col: Last column number.
        """
        top_side = Side(style="thin", color="111827")
        for row in rows:
            for col in range(start_col, end_col + 1):
                cell = ws.cell(row=row, column=col)
                cell.font = Font(bold=True, color=cell.font.color.rgb if cell.font.color and cell.font.color.type == "rgb" else "000000")
                cell.border = Border(top=top_side, bottom=self.thin_gray)
